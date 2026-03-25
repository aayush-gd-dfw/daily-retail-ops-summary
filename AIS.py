import os
import io
import re
import base64
from datetime import datetime, date, timezone
from typing import Optional, Dict, Any, List, Tuple

import requests
from msal import ConfidentialClientApplication
from openpyxl import load_workbook


GRAPH = "https://graph.microsoft.com/v1.0"

MAILBOX_UPN = os.getenv("MAILBOX_UPN", "apatil@glassdoctordfw.com")

SUBJECT_COMPLETED = "11 PM Completed Revenue"
SUBJECT_UPCOMING = "11 PM Upcoming Revenue"
SUBJECT_JOB_NOTES = "11 PM Job Notes"

SHAREPOINT_SHEET_NAME = "Daily Summary"


# -------------------- Small helpers --------------------
def must_env(name: str) -> str:
    v = os.getenv(name)
    if not v:
        raise RuntimeError(f"Missing environment variable: {name}")
    return v


def parse_dt(dt_str: str) -> datetime:
    if not dt_str:
        return datetime(1970, 1, 1, tzinfo=timezone.utc)
    if dt_str.endswith("Z"):
        dt_str = dt_str.replace("Z", "+00:00")
    return datetime.fromisoformat(dt_str)


def parse_money(val) -> float:
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = re.sub(r"[^0-9.\-]", "", str(val))
    return float(s) if s else 0.0


def normalize_job(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if not s:
        return ""
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    return s


def try_parse_any_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if v is None:
        return None

    s = str(v).strip()
    if not s:
        return None

    for fmt in (
        "%m/%d/%Y", "%m/%d/%y",
        "%Y-%m-%d", "%Y/%m/%d",
        "%m-%d-%Y", "%m-%d-%y"
    ):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass

    try:
        return datetime.fromisoformat(s).date()
    except Exception:
        return None


def fmt_money(n: float) -> str:
    return f"${n:,.2f}"


def find_col_idx(header, target_names_lower):
    h = [str(x).strip().lower() for x in header]
    for i, name in enumerate(h):
        if name in target_names_lower:
            return i
    for i, name in enumerate(h):
        for t in target_names_lower:
            if t in name:
                return i
    return None


def extract_first_date_from_filename(fname: str) -> date:
    """
    Example:
    11 PM Completed Revenue 03_24_26 - 03_24_26.xlsx -> 2026-03-24
    """
    if not fname:
        raise ValueError("Missing filename for date extraction.")

    m = re.search(r"(\d{1,2})[._-](\d{1,2})[._-](\d{2,4})", fname)
    if not m:
        raise ValueError(f"Could not extract date from filename: {fname}")

    mm, dd, yy = map(int, m.groups())
    yyyy = 2000 + yy if yy < 100 else yy
    return date(yyyy, mm, dd)


# -------------------- Auth --------------------
def get_token() -> str:
    tenant_id = must_env("TENANT_ID")
    client_id = must_env("CLIENT_ID")
    client_secret = must_env("CLIENT_SECRET")

    app = ConfidentialClientApplication(
        client_id=client_id,
        client_credential=client_secret,
        authority=f"https://login.microsoftonline.com/{tenant_id}",
    )

    result = app.acquire_token_for_client(
        scopes=["https://graph.microsoft.com/.default"]
    )

    if "access_token" not in result:
        raise RuntimeError(f"Token error: {result.get('error')} - {result.get('error_description')}")

    return result["access_token"]


def graph_get(token: str, url: str, params: Optional[dict] = None) -> Dict[str, Any]:
    headers = {
        "Authorization": f"Bearer {token}",
        "ConsistencyLevel": "eventual",
    }
    r = requests.get(url, headers=headers, params=params, timeout=60)
    if not r.ok:
        try:
            print("Graph error payload:", r.json())
        except Exception:
            print("Graph error text:", r.text)
        r.raise_for_status()
    return r.json()


def graph_get_bytes(token: str, url: str, params: Optional[dict] = None) -> bytes:
    headers = {
        "Authorization": f"Bearer {token}",
        "ConsistencyLevel": "eventual",
    }
    r = requests.get(url, headers=headers, params=params, timeout=120)
    if not r.ok:
        try:
            print("Graph error payload:", r.json())
        except Exception:
            print("Graph error text:", r.text)
        r.raise_for_status()
    return r.content


def graph_put_bytes(token: str, url: str, content: bytes, content_type: str) -> Dict[str, Any]:
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": content_type,
    }
    r = requests.put(url, headers=headers, data=content, timeout=180)
    if not r.ok:
        try:
            print("Graph error payload:", r.json())
        except Exception:
            print("Graph error text:", r.text)
        r.raise_for_status()
    return r.json()


# -------------------- Outlook --------------------
def messages_for_subject(token: str, mailbox_upn: str, subject_phrase: str, top_n: int = 25) -> List[Dict[str, Any]]:
    """
    Uses $search and sorts locally.
    """
    url = f"{GRAPH}/users/{mailbox_upn}/mailFolders/Inbox/messages"
    params = {
        "$select": "id,subject,receivedDateTime,from,hasAttachments",
        "$top": str(top_n),
        "$search": f"\"{subject_phrase}\"",
    }

    data = graph_get(token, url, params=params)
    msgs = data.get("value", [])

    phrase = subject_phrase.lower()
    candidates = [m for m in msgs if phrase in (m.get("subject") or "").lower()]
    candidates.sort(key=lambda m: parse_dt(m.get("receivedDateTime", "")), reverse=True)
    return candidates


def get_first_xlsx_attachment_from_message(token: str, mailbox_upn: str, message_id: str) -> Tuple[Optional[str], Optional[bytes]]:
    url = f"{GRAPH}/users/{mailbox_upn}/messages/{message_id}/attachments"
    data = graph_get(token, url, params={"$top": "50"})
    atts = data.get("value", [])

    for a in atts:
        name = (a.get("name") or "")
        if name.lower().endswith(".xlsx"):
            cb = a.get("contentBytes")
            if cb:
                return name, base64.b64decode(cb)

            att_id = a.get("id")
            if att_id:
                raw_url = f"{GRAPH}/users/{mailbox_upn}/messages/{message_id}/attachments/{att_id}/$value"
                b = graph_get_bytes(token, raw_url)
                return name, b

    return None, None


# -------------------- SharePoint workbook --------------------
def download_sharepoint_excel(token: str, drive_id: str, item_id: str) -> bytes:
    url = f"{GRAPH}/drives/{drive_id}/items/{item_id}/content"
    return graph_get_bytes(token, url)


def upload_sharepoint_excel(token: str, drive_id: str, item_id: str, content: bytes) -> None:
    url = f"{GRAPH}/drives/{drive_id}/items/{item_id}/content"
    graph_put_bytes(
        token,
        url,
        content,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# -------------------- Excel parsing --------------------
def read_xlsx_first_sheet_rows(xlsx_bytes: bytes) -> List[List[Any]]:
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for r in ws.iter_rows(values_only=True):
        rows.append([("" if v is None else v) for v in r])
    return rows


def parse_completed_report(rows: List[List[Any]]) -> Tuple[Dict[str, Dict[str, Any]], float]:
    """
    Completed report:
    - Invoice # = job number
    - Jobs Subtotal = revenue
    - last summary row carries total completed revenue
    """
    if not rows or len(rows) < 2:
        raise ValueError("Completed report is empty or missing rows.")

    header = rows[0]
    body = rows[1:]

    invoice_col = find_col_idx(header, {"invoice #", "invoice"})
    subtotal_col = find_col_idx(header, {"jobs subtotal", "subtotal"})
    bu_col = find_col_idx(header, {"business unit"})

    if invoice_col is None or subtotal_col is None:
        raise ValueError(f"Could not find required columns in Completed report. Header: {header}")

    jobs = {}
    completed_revenue_total = 0.0

    for r in body:
        if len(r) <= max(invoice_col, subtotal_col):
            continue

        invoice_val = normalize_job(r[invoice_col])
        subtotal_val = parse_money(r[subtotal_col])

        bu_val = ""
        if bu_col is not None and len(r) > bu_col and r[bu_col] is not None:
            bu_val = str(r[bu_col]).strip()

        # total row
        if not invoice_val and subtotal_val and (bu_col is None or not bu_val):
            completed_revenue_total = subtotal_val
            continue

        if invoice_val:
            jobs[invoice_val] = {
                "revenue": subtotal_val
            }

    # fallback in case total row is not found
    if completed_revenue_total == 0.0:
        completed_revenue_total = round(sum(v["revenue"] for v in jobs.values()), 2)

    return jobs, round(completed_revenue_total, 2)


def build_jobs_from_upcoming_for_target_date(rows: List[List[Any]], target_date: date) -> Dict[str, Dict[str, Any]]:
    """
    For yesterday upcoming:
    get only jobs where Next Appt Start Date == target_date
    """
    if not rows or len(rows) < 3:
        raise ValueError("Upcoming report is empty or missing rows.")

    header = rows[0]
    body = rows[2:]  # skip 2nd row like "Next Appt Start Date: 3/24/2026"

    job_col = find_col_idx(header, {"job #", "job number", "job"})
    date_col = find_col_idx(header, {"next appt start date", "appointment start date", "next appt"})
    rev_col = find_col_idx(header, {"jobs subtotal", "subtotal"})

    if job_col is None or date_col is None or rev_col is None:
        raise ValueError(f"Could not find required columns in Upcoming report. Header: {header}")

    jobs = {}

    for r in body:
        if len(r) <= max(job_col, date_col, rev_col):
            continue

        d = try_parse_any_date(r[date_col])
        if d != target_date:
            continue

        job = normalize_job(r[job_col])
        if not job:
            continue

        jobs[job] = {
            "revenue": parse_money(r[rev_col])
        }

    return jobs


def build_all_jobs_from_upcoming(rows: List[List[Any]]) -> Dict[str, Dict[str, Any]]:
    """
    For today's upcoming:
    scan entire file for missing job number regardless of date
    """
    if not rows or len(rows) < 3:
        raise ValueError("Upcoming report is empty or missing rows.")

    header = rows[0]
    body = rows[2:]  # skip date-label row

    job_col = find_col_idx(header, {"job #", "job number", "job"})
    rev_col = find_col_idx(header, {"jobs subtotal", "subtotal"})

    if job_col is None or rev_col is None:
        raise ValueError(f"Could not find required columns in Upcoming report. Header: {header}")

    jobs = {}
    for r in body:
        if len(r) <= max(job_col, rev_col):
            continue

        job = normalize_job(r[job_col])
        if not job:
            continue

        jobs[job] = {
            "revenue": parse_money(r[rev_col])
        }

    return jobs


def build_job_notes_lookup(rows: List[List[Any]]) -> Dict[str, Dict[str, Any]]:
    if not rows or len(rows) < 2:
        raise ValueError("Job Notes report is empty or missing rows.")

    header = rows[0]
    body = rows[1:]

    job_col = find_col_idx(header, {"jobnumber", "job number", "job #", "job"})
    note_col = find_col_idx(header, {"jobnote", "job note", "job notes", "note", "notes"})
    rev_col = find_col_idx(header, {"jobs subtotal", "subtotal"})

    if job_col is None or note_col is None:
        raise ValueError(f"Could not find required columns in Job Notes report. Header: {header}")

    notes = {}
    for r in body:
        if len(r) <= max(job_col, note_col):
            continue

        job = normalize_job(r[job_col])
        note = str(r[note_col]).strip() if r[note_col] is not None else ""
        revenue = parse_money(r[rev_col]) if rev_col is not None and len(r) > rev_col else 0.0

        if job and note:
            notes[job] = {
                "note": note,
                "revenue": revenue
            }

    return notes


# -------------------- Analysis --------------------
def analyze_reports(
    completed_fname: str,
    completed_bytes: bytes,
    upcoming_today_bytes: bytes,
    upcoming_yesterday_bytes: bytes,
    notes_bytes: bytes,
) -> Dict[str, Any]:

    target_date = extract_first_date_from_filename(completed_fname)

    completed_rows = read_xlsx_first_sheet_rows(completed_bytes)
    upcoming_today_rows = read_xlsx_first_sheet_rows(upcoming_today_bytes)
    upcoming_yesterday_rows = read_xlsx_first_sheet_rows(upcoming_yesterday_bytes)
    notes_rows = read_xlsx_first_sheet_rows(notes_bytes)

    completed_jobs, completed_revenue = parse_completed_report(completed_rows)
    upcoming_today_all_jobs = build_all_jobs_from_upcoming(upcoming_today_rows)
    scheduled_yesterday_for_today = build_jobs_from_upcoming_for_target_date(upcoming_yesterday_rows, target_date)
    job_notes_lookup = build_job_notes_lookup(notes_rows)

    scheduled_ids = set(scheduled_yesterday_for_today.keys())
    completed_ids = set(completed_jobs.keys())

    missing_jobs = scheduled_ids - completed_ids
    extra_jobs = completed_ids - scheduled_ids

    missing_rescheduled_list = []
    jobs_with_notes_list = []

    for job in sorted(missing_jobs):
        if job in upcoming_today_all_jobs:
            missing_rescheduled_list.append(
                f"{job} - {fmt_money(upcoming_today_all_jobs[job]['revenue'])}"
            )
        elif job in job_notes_lookup:
            note_text = job_notes_lookup[job]["note"] if isinstance(job_notes_lookup[job], dict) else job_notes_lookup[job]
        
            # prefer today's upcoming revenue if available, otherwise yesterday scheduled revenue
            if job in upcoming_today_all_jobs:
                note_revenue = upcoming_today_all_jobs[job]["revenue"]
            else:
                note_revenue = scheduled_yesterday_for_today.get(job, {}).get("revenue", 0.0)
        
            if note_text:
                jobs_with_notes_list.append(
                    f'{job} - "{note_text}" - {fmt_money(note_revenue)}'
                )
            else:
                jobs_with_notes_list.append(
                    f"{job} - {fmt_money(note_revenue)}"
                )

    extra_jobs_list = [
        f"{job} - {fmt_money(completed_jobs[job]['revenue'])}"
        for job in sorted(extra_jobs)
    ]

    scheduled_revenue = round(
        sum(v["revenue"] for v in scheduled_yesterday_for_today.values()), 2
    )

    return {
        "date": target_date,
        "scheduled_revenue": scheduled_revenue,
        "completed_revenue": completed_revenue,
        "missing_jobs_rescheduled": ", ".join(missing_rescheduled_list),
        "jobs_with_notes": ", ".join(jobs_with_notes_list),
        "jobs_not_scheduled_but_happened": ", ".join(extra_jobs_list),
        "scheduled_count": len(scheduled_ids),
        "completed_count": len(completed_ids),
        "missing_count": len(missing_jobs),
        "extra_count": len(extra_jobs),
    }


# -------------------- Daily Summary write --------------------
def ensure_daily_summary_sheet(wb):
    if SHAREPOINT_SHEET_NAME in wb.sheetnames:
        return wb[SHAREPOINT_SHEET_NAME]

    ws = wb.create_sheet(SHAREPOINT_SHEET_NAME)
    ws.append([
        "Date",
        "Scheduled Revenue",
        "Completed Revenue",
        "Missing jobs that are rescheduled",
        "Jobs with notes",
        "Jobs which were not scheduled for today but happened today",
    ])
    return ws


def find_existing_row_by_date(ws, target_date: date) -> Optional[int]:
    for row in range(2, ws.max_row + 1):
        v = ws.cell(row=row, column=1).value
        d = try_parse_any_date(v)
        if d == target_date:
            return row
    return None


def write_daily_summary_to_workbook(wb, summary: Dict[str, Any]) -> int:
    ws = ensure_daily_summary_sheet(wb)

    existing_row = find_existing_row_by_date(ws, summary["date"])
    target_row = existing_row if existing_row else ws.max_row + 1

    ws.cell(row=target_row, column=1).value = summary["date"].strftime("%m/%d/%Y")
    ws.cell(row=target_row, column=2).value = summary["scheduled_revenue"]
    ws.cell(row=target_row, column=3).value = summary["completed_revenue"]
    ws.cell(row=target_row, column=4).value = summary["missing_jobs_rescheduled"]
    ws.cell(row=target_row, column=5).value = summary["jobs_with_notes"]
    ws.cell(row=target_row, column=6).value = summary["jobs_not_scheduled_but_happened"]

    return target_row


# -------------------- Main --------------------
def main():
    token = get_token()

    drive_id = must_env("DRIVE_ID")
    file_item_id = must_env("FILE_ITEM_ID")

    completed_msgs = messages_for_subject(token, MAILBOX_UPN, SUBJECT_COMPLETED, top_n=10)
    upcoming_msgs = messages_for_subject(token, MAILBOX_UPN, SUBJECT_UPCOMING, top_n=10)
    job_notes_msgs = messages_for_subject(token, MAILBOX_UPN, SUBJECT_JOB_NOTES, top_n=10)

    if not completed_msgs:
        raise RuntimeError(f"No email found for subject: {SUBJECT_COMPLETED}")

    if len(upcoming_msgs) < 2:
        raise RuntimeError(f"Need at least 2 emails for subject: {SUBJECT_UPCOMING}")

    if not job_notes_msgs:
        raise RuntimeError(f"No email found for subject: {SUBJECT_JOB_NOTES}")

    completed_msg = completed_msgs[0]
    upcoming_today_msg = upcoming_msgs[0]
    upcoming_yesterday_msg = upcoming_msgs[1]
    job_notes_msg = job_notes_msgs[0]

    print("Using emails:")
    print("Completed:", completed_msg.get("subject"), completed_msg.get("receivedDateTime"))
    print("Upcoming Today:", upcoming_today_msg.get("subject"), upcoming_today_msg.get("receivedDateTime"))
    print("Upcoming Yesterday:", upcoming_yesterday_msg.get("subject"), upcoming_yesterday_msg.get("receivedDateTime"))
    print("Job Notes:", job_notes_msg.get("subject"), job_notes_msg.get("receivedDateTime"))

    completed_fname, completed_bytes = get_first_xlsx_attachment_from_message(
        token, MAILBOX_UPN, completed_msg["id"]
    )
    upcoming_today_fname, upcoming_today_bytes = get_first_xlsx_attachment_from_message(
        token, MAILBOX_UPN, upcoming_today_msg["id"]
    )
    upcoming_yesterday_fname, upcoming_yesterday_bytes = get_first_xlsx_attachment_from_message(
        token, MAILBOX_UPN, upcoming_yesterday_msg["id"]
    )
    job_notes_fname, job_notes_bytes = get_first_xlsx_attachment_from_message(
        token, MAILBOX_UPN, job_notes_msg["id"]
    )

    if not completed_bytes:
        raise RuntimeError("Completed report attachment not found.")
    if not upcoming_today_bytes:
        raise RuntimeError("Today's upcoming report attachment not found.")
    if not upcoming_yesterday_bytes:
        raise RuntimeError("Yesterday's upcoming report attachment not found.")
    if not job_notes_bytes:
        raise RuntimeError("Job Notes report attachment not found.")

    print("Attachments found:")
    print("Completed file:", completed_fname)
    print("Upcoming today file:", upcoming_today_fname)
    print("Upcoming yesterday file:", upcoming_yesterday_fname)
    print("Job notes file:", job_notes_fname)

    summary = analyze_reports(
        completed_fname=completed_fname,
        completed_bytes=completed_bytes,
        upcoming_today_bytes=upcoming_today_bytes,
        upcoming_yesterday_bytes=upcoming_yesterday_bytes,
        notes_bytes=job_notes_bytes,
    )

    print("\nSummary generated:")
    for k, v in summary.items():
        print(f"{k}: {v}")

    xls_bytes = download_sharepoint_excel(token, drive_id, file_item_id)
    wb = load_workbook(io.BytesIO(xls_bytes))

    row_num = write_daily_summary_to_workbook(wb, summary)

    out = io.BytesIO()
    wb.save(out)
    upload_sharepoint_excel(token, drive_id, file_item_id, out.getvalue())

    print(f"\nDone. Daily Summary updated on row {row_num} and uploaded back to SharePoint.")


if __name__ == "__main__":
    main()
